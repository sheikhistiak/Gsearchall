import openpyxl as openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import time
from datetime import datetime


def search_all(select_sheet):

    global Last_item
    global First_item

    def load_data(file):
        workbook = openpyxl.load_workbook(file)

        worksheet = workbook[select_sheet]

        # Iterate through the rows and collect values from column C (skipping first two values)
        c_column_values = []
        for index, row in enumerate(worksheet.iter_rows(min_col=3, max_col=3)):
            if index >= 2:
                c_column_values.append(row[0].value)

        workbook.close()
        # Print the collected values
        # print("Values in column C of", current_worksheet)
        # print(c_column_values)
        return c_column_values
        # Close the workbook


    # Load the Excel file
    # file_path = 'Excel.xlsx'  # Replace with your file path
    load_data(file_path)

    def perform_google_search(keyword):
        # Create a new instance of the Chrome driver
        driver = webdriver.Chrome()

        # Navigate to the Google site
        driver.get("https://www.google.com/")

        # Wait for the site to load
        driver.implicitly_wait(10)

        # Find the search box element and enter keywords
        search_box = driver.find_element(By.ID, 'APjFqb')
        search_box.send_keys(keyword)

        # Wait for a few seconds (you might need to adjust this)
        time.sleep(5)

        # Perform the search by pressing the 'RETURN' key
        # search_box.send_keys(Keys.RETURN)
        # search_box.send_keys(Keys.SPACE)

        # Wait for the search results to load
        driver.implicitly_wait(10)

        # Get the HTML source of the search results page
        html = driver.page_source

        # Parse the HTML using BeautifulSoup
        soup = BeautifulSoup(html, 'html.parser')
        soup_pretty = soup.prettify()

        # Print the prettified HTML to the console
        # print(soup_pretty)

        # Find all the search list on the page
        Search_list = soup.find_all('div', class_='wM6W7d')
        Search_list_str = str(Search_list)
        # print(Search_list_str)

        # Wait for a few seconds
        time.sleep(5)

        # Update the soup
        html = driver.page_source

        # Parse the HTML using BeautifulSoup

        # Parse the input data using BeautifulSoup
        soup = BeautifulSoup(Search_list_str, 'html.parser')

        # Find all the search list on the page
        search_results = soup.find_all('div', class_='wM6W7d')

        # Extract text from each span and store in a list
        text_list = []
        for search_result in search_results:
            span_element = search_result.find('span')
            if span_element:
                text = span_element.text.strip()  # Extract text and remove leading/trailing spaces
                if text:
                    text_list.append(text)

        # List the extracted text in ascending order based on the length of the search results
        text_list.sort(key=len)

        # Print the rearranged text list
        # for item in text_list:
        # print(item)

        # Return the rearranged text list
        result_list = text_list
        # print(result_list)
        driver.quit()
        return result_list
        # Close the browser window


    # List of keywords to search
    keywords = load_data(file_path)

    # Start iterating from the 3rd row
    start_row = 3

    # Call the function for each keyword and print the result
    for keyword in keywords:
        google_results = perform_google_search(keyword)

        # Calculate the length of the list
        list_length = len(google_results)

        # Print the first and last items based on the length
        if list_length > 0:
            First_item = google_results[0]
            # print(First_item)
            Last_item = google_results[list_length - 1]
            # print(Last_item)
        else:
            print("The list is empty.")

        workbook = openpyxl.load_workbook(file_path)

        worksheet = workbook[select_sheet]

        print(worksheet)

        # Get input for values in columns D and E
        value_d = Last_item
        value_e = First_item
        
        # Write the values into the worksheet
        worksheet.cell(row=start_row, column=4, value=value_d)
        worksheet.cell(row=start_row, column=5, value=value_e)

        start_row += 1

        # Save the changes to the Excel file
        workbook.save(file_path)
        # print(worksheet)

def read_load_data(file_path):
    workbook = openpyxl.load_workbook(file_path)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Sheet: {sheet_name}")
        print(sheet_name)
        search_all(sheet_name)
        print("\n")
        print("\n")


# Replace with the path to your existing Excel file
file_path = "Excel.xlsx"
read_load_data(file_path)
